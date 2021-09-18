import requests
import json
import pandas as pd
import os.path
from warnings import warn
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import seaborn as sns
import re
import pickle
import numpy as np


class SessionWithUrlBase(requests.Session):
    """
    from https://github.com/quintel/third-party/blob/master/Python_ETM_API/ETM_API.py

    Helper class to store the base url. This allows us to only type the
    relevant additional information.
    from: https://stackoverflow.com/questions/42601812/python-requests-url-base-in-session
    """

    def __init__(self, url_base=None, *args, **kwargs):

        super(SessionWithUrlBase, self).__init__(*args, **kwargs)

        if url_base is None:
            url_base = "https://engine.energytransitionmodel.com/api/v3"

        self.url_base = url_base

    def request(self, method, url, **kwargs):
        modified_url = self.url_base + url

        return super(SessionWithUrlBase, self).request(method, modified_url, **kwargs)


# helper function to get rid of most html scraping artefacts
def cleanhtml(raw_html):
    if isinstance(raw_html, str):
        cleanr = re.compile("<.*?>")
        cleanertext = re.sub(cleanr, "", raw_html)
        cleanr2 = re.compile("&([a-z0-9]+|#[0-9]{1,6}|#x[0-9a-f]{1,6});")
        cleanerdertext = re.sub(cleanr2, "", cleanertext)
        cleanr3 = re.compile("\\r")
        cleantext = re.sub(cleanr3, "", cleanerdertext)
        cleantext = cleantext.strip().replace("\n", "")
    else:
        cleantext = raw_html
    return cleantext


# helper function to convert cols to numbers
def col_to_num(col_str):
    """
    Convert base26 column string to number.
        addapted from: https://stackoverflow.com/questions/7261936/
    """
    expn = 0
    col_num = -1  # change to obtain index locations instead of col numbers
    for char in reversed(col_str):
        col_num += (ord(char) - ord("A") + 1) * (26 ** expn)
        expn += 1

    return col_num


class ETMapi:
    def __init__(
        self,
        area_code="nl",
        scenario_id=None,
        end_year=2050,
        source="api-WB",
        url_base=None,
    ):

        self.session = SessionWithUrlBase(url_base=url_base)
        self.source = source
        self.area_code = area_code
        self.end_year = end_year
        self.browse_url = "https://pro.energytransitionmodel.com/scenarios/"
        self.headers = {
            "Content-Type": "application/json",
        }
        self.verbose = True
        self.debug = False
        self.user_values = dict()
        # some default metrics to query
        self.gqueries = [
            "dashboard_reduction_of_co2_emissions_versus_1990",
            "dashboard_renewability",
            "dashboard_share_of_renewable_electricity",
            "dashboard_energy_import_netto",
            "dashboard_total_costs",
            "dashboard_energy_demand_primary_of_final_plus_export_losses",
            "dashboard_blackout_hours",
            "dashboard_total_number_of_excess_events",
            "dashboard_security_of_supply",
            "lv_net_capacity_delta_present_future",
            "mv_net_capacity_delta_present_future",
            "hv_net_capacity_delta_present_future",
        ]

        self.scenario_id = scenario_id

    @property
    def scenario_id(self):
        return self._scenario_id

    @scenario_id.setter
    def scenario_id(self, value):
        self._scenario_id = self.id_extractor(value)

    def get_title(self, scenario_id):

        if scenario_id is not None:
            r = self.session.get(f"/scenarios/{scenario_id}/")
            if r.status_code != 200:
                raise ValueError("Response not 200")
            di = json.loads(r.content)
            title = di["title"]
        else:
            title = ""

        return title

    def create_new_scenarios_from_excel(
        self,
        file_name,
        start_row,
        end_row,
        ref_scenarios,
        end_years,
        sheet_name=None,
        start_col=None,
        end_col=None,
        scenario_cols=None,
        titles=None,
    ):
        """
        Documentation missing @seth
        """
        # if needed convert ref_scenarios/end_years to the same length as amount of scenarios
        if scenario_cols is not None:
            nscenarios = len(scenario_cols)
        elif end_col is not None and start_col is not None:
            nscenarios = col_to_num(end_col) - col_to_num(start_col) + 1

        if not isinstance(end_years, list):
            end_years = list([end_years]) * nscenarios
        if not isinstance(ref_scenarios, list):
            ref_scenarios = list([ref_scenarios]) * nscenarios
        if not isinstance(titles, list):
            titles = list([titles]) * nscenarios

        all_user_values = self.extract_excel_etm_values(
            file_name=file_name,
            start_row=start_row,
            end_row=end_row,
            sheet_name=sheet_name,
            start_col=start_col,
            end_col=end_col,
            scenario_cols=scenario_cols,
        )

        results = dict()
        for i, key in enumerate(all_user_values.keys()):

            self.user_values = all_user_values[key]
            self.end_year = end_years[i]
            self.scenario_id = ref_scenarios[i]
            # order matters (scenario_id setter overwrites the self.title) @seth
            if titles is not None:
                self.title = titles[i]
            else:
                self.title = key
            self.create_new_scenario(title=self.title, use_custom_values=True)
            results.update({key: {"metrics": self.metrics, "url": self.browse_url}})

        self.scenario_batch_result = results

    @staticmethod
    def extract_excel_etm_values(
        file_name,
        start_row,
        end_row,
        sheet_name=None,
        start_col=None,
        end_col=None,
        scenario_cols=None,
    ):
        """
        Documentation missing @seth
        """
        if sheet_name is None:
            sheet_name = "Sheet1"

        # read sheet
        df = pd.read_excel(file_name, sheet_name=sheet_name, engine="openpyxl")

        # find start and en row in pandas indices
        start_row = start_row - 2
        end_row = end_row - 1

        if scenario_cols is not None:

            # convert cols to pandas indices
            scenario_nums = [col_to_num(col) for col in scenario_cols]
            # get scenario data
            scenarios = df.iloc[start_row:end_row, scenario_nums]

        elif end_col is not None and start_col is not None:

            # convert cols to pandas indices
            start = col_to_num(start_col)
            end = col_to_num(end_col) + 1
            # get scenario data
            scenarios = df.iloc[start_row:end_row, start:end]

        scenarios.columns = [cleanhtml(col) for col in scenarios.columns]

        # get translated names as column
        scenarios["trans_name"] = (
            df.iloc[start_row:end_row, 0]
            + df.iloc[start_row:end_row, 1]
            + df.iloc[start_row:end_row, 2].fillna("")
        ).values

        # open the key map
        with open("data/name_to_key.pkl", "rb") as handle:
            name_to_key = pickle.load(handle)
            keys = list(name_to_key.values())

        # name-to-key function
        def n2k(index):
            # get trans_name
            name = scenarios.trans_name[index]
            key = name_to_key.get(name, None)
            # if no exact match, revert to index based (could be wrong tho)
            if key is None:
                key = keys[index]
                warn(
                    f"No key match, reverting to index based assumption on row {index}, found key {key}."
                )
            return key

        all_user_values = dict()
        for col in scenarios.columns:
            if col != "trans_name":
                scenario = scenarios[col]
                user_values = {
                    n2k(index): row
                    for index, row in scenario.iteritems()
                    if not np.isnan(row)
                }
                all_user_values.update({col: user_values})

        return all_user_values

    def create_new_scenario(self, title, use_custom_values=False):
        """
        POST-method to create a scenario in ETM using ETM v3 API. Converts non-strings to strings.
        Pushes the setup as json in correct format voor ETM v3 API.

        arguments:
            area_code (optional, one of the existing area codes (use etm.area_codes() to obtain))
            title (title of scenario)
            end_year (interger end year)
            source (paper trail, recommended to use)
            scenario_id (refference scenario)
            verbose (false turns off all print and warn)

        returns:
            requests response object
        """

        scenario_setup = {
            "scenario": {
                "area_code": str(self.area_code),
                "title": str(title),
                "end_year": str(self.end_year),
                "source": str(self.source),
                "scenario_id": str(self.scenario_id),
            }
        }

        url = "/scenarios/"
        r = self.session.post(url, json=scenario_setup, headers=self.headers)

        if r.status_code == 200:

            api_url = json.loads(r.content)["url"]
            self.scenario_id = api_url.split("/")[-1]
            self.browse_url += self.scenario_id

            if use_custom_values:
                p = self._change_inputs()
                if not p.status_code == 200:
                    raise ValueError(f"Response not succesful: {p.json()['errors']}")
                self.update_metrics()

            if self.verbose:
                print()
                print("Browsable URL to scenario:")
                print(self.browse_url)

        if not r.status_code == 200:
            raise ValueError(f"Response not succesful: {r.status_code}")

        if self.debug:
            return r
        else:
            pass

    def update_metrics(self, output_format="dataframe"):
        """
        Put querries and store results to ETM bases on self.gqueries
        """
        input_data = {
            "gqueries": self.gqueries,
        }

        url = "/scenarios/"
        p = self.session.put(
            url + str(self.scenario_id), json=input_data, headers=self.headers
        )
        if p.status_code != 200:
            warn(f"Gquerry not succesful for {self.scenario_id}")
            print("p")

        # store metrics
        if output_format == "dataframe":
            self.metrics = pd.DataFrame.from_dict(p.json()["gqueries"], orient="index")
        elif output_format == "dict":
            self.metrics = p.json()["gqueries"]
        else:
            self.metrics = p.json()["gqueries"]

        pass

    def _change_inputs(self):
        """
        Change inputs to ETM bases on self.user_values.
        """
        input_data = {
            "scenario": {"user_values": self.user_values},
            "detailed": True,
        }

        url = "/scenarios/"
        p = self.session.put(
            url + self.scenario_id, json=input_data, headers=self.headers
        )
        
        if p.status_code != 200:
            raise ValueError(f"Response not succesful: {p.json()['errors']}")

        return p

    def update_inputs(self, debug=False):
        """Basicly ._change_inputs but does not return p out of debug"""
        p = self._change_inputs()
        if debug:
            return p
        else:
            pass

    def get_areas(self, filepath=None, refresh=False, save_csv=True):
        """
        Use ETM v3 API to request all areas and from that extract only the area codes, to use in
        other ETM v3 API reuqests.

        arguments:
            filepath:
                specify location to save to / read from (default: /etm_ref/)
            refresh:
                specify true to force refresh, checks for existing file on path location to prevent
                errors from trying to load non-existing files
            save_cvs:
                caches the api response to a csv to speed up the process if called again.

        returns
            set with areas (as atribute)
            (saves csv to save requests and allow browsing of codes)
        """

        if filepath is None:
            filepath = "data/areacodes.pkl"

        if refresh or not os.path.isfile(filepath):

            url = "/areas/"
            r = self.session.get(url)
            r_list = json.loads(r.content)
            self.areas_raw = r_list

            # create a set to refer to if needed (drops dupes & sorts)
            areas = set([item["area"] for item in r_list if item["useable"]])

            df = pd.DataFrame(r_list)

            if save_csv:
                df.to_pickle(filepath)

        else:
            df = pd.read_pickle(filepath)
            areas = set(df["area"])

        self.areas = areas
        return areas

    def get_area_settings(area_code):
        """
        Gets the area settings based on the area code.

        argument:
            area_code (as defined by ETM)

        returns:
            requests response object
        """

        request_url = (
            f"https://engine.energytransitionmodel.com/api/v3/areas/{area_code}"
        )
        response = requests.get(request_url)

        return response

    def generate_input_worksheet(
        self, filepath=None, scenario_list=None, prettify=False
    ):
        """
        Method to generate an empty input worksheet. Can later be used to upload new scenarios
        and very useful for documenting all changes and their sources.

        Arguments:
            filepath: filepath of output worksheet
            scenario_list: list of scenario_ids to use as templates, included as column in sheet
            prettify: removes keys (non-translated identification labels) and formats the worksheet

        returns:
            none, creates a file on specified location
        """

        if scenario_list is None:
            scenario_list = [self.scenario_id]

        if filepath is None:
            filepath = "latest_generated_worksheet.xlsx"
            if self.scenario_id is not None:
                filepath = f"latest_generated_worksheet_{self.scenario_id}.xlsx"

        # load in front-end variables based on scraped contents
        ref = pd.read_pickle("data/inputs_reference.pkl")

        # take only what we are interested in
        df = ref[["key", "group", "subgroup", "translated_name", "unit"]].copy()
        df.columns = ["key", "Section", "Subsection", "Slider name", "Unit"]

        for sid in scenario_list:

            self.scenario_id = sid  # use setter
            if self.scenario_id is not None:
                col_name = self.title + f" ({self.scenario_id})"
            else:
                col_name = "Default scenario"

            df.loc[:, col_name] = pd.Series(dtype="float64")
            # request the inputs based on scenario id (if none --> default)
            scenario = self.get_all_inputs(
                outputformat="dataframe", scenario_id=self.scenario_id
            )

            # look up the values in the front end ETM (based on scraped list)
            user = "user" in scenario.index

            for index, key in df.key.iteritems():

                # add all default values
                df.loc[index, col_name] = scenario.loc["default", key]

                # overwrite if user specified values
                if user:

                    value = scenario.loc["user", key]
                    if not pd.isna(value):
                        df.loc[index, col_name] = value

        # apply cleaner function and assign to df
        df.loc[:, "Slider description"] = ref["sanitized_description"].apply(cleanhtml)

        # write out
        if prettify:
            df.drop(labels="key", axis=1, inplace=True)
        df.to_excel(filepath, index=False, engine="openpyxl")

        if prettify:

            n_colors = len(df["Section"].unique())
            pastel = sns.color_palette("pastel", n_colors=n_colors)
            pastel = pastel.as_hex()
            muted = sns.color_palette("muted", n_colors=n_colors)
            muted = muted.as_hex()

            section_number = lambda key: list(df.Section.unique()).index(key)
            subsection_number = lambda key: list(df.Subsection.unique()).index(key)
            tuple_gen = lambda row: (
                section_number(row.Section),
                subsection_number(row.Subsection),
            )
            color_flipper = lambda x: pastel[x[0]] if x[1] % 2 else muted[x[0]]

            colors = df.apply(tuple_gen, axis=1).apply(color_flipper)

            wb = load_workbook(filepath)
            ws = wb[wb.sheetnames[0]]

            for index, cell in enumerate(ws["A"]):
                if index != 0:
                    fill = PatternFill(
                        fill_type="solid",
                        start_color=colors[index - 1].replace("#", "FF"),
                    )
                    ws["A"][index].fill = fill
                    ws["B"][index].fill = fill
                    ws["C"][index].fill = fill
                    ws["D"][index].fill = fill

            wb.save(filepath)

        pass

    def get_all_inputs(self, outputformat="list", scenario_id=None):
        """
        Fetch all possible inputs (this is more than available in the front-end)
        @seth implement a work-around to always get the default scenario
        """
        if scenario_id is not None:
            url = f"/scenarios/{scenario_id}/inputs/"
        else:
            url = "/inputs/"

        r = self.session.get(url)

        if outputformat == "list":
            inputs = list(json.loads(r.content).keys())
        elif outputformat == "dataframe":
            inputs = pd.DataFrame(r.json())
        elif outputformat == "response":
            inputs = r

        if not r.status_code == 200:
            warn(f"Response not succesful: {r.status_code}")

        return inputs

    def id_extractor(self, scenario_id):
        """
        Basicly a lengthy way to get a session ID or the original uploading ID for extracting the default
        values of any scenario. Without this workaround it is not possible to extract values of saved_templates,
        as it is not (yet) implemented in the ETM. Uses bs4.

        argument
            scenario_id: 4 or 6 digit code of the scenario to extract

        returns
            _scenario_id for assingment to self
        """

        # if 4 digit and not None (also len = 4) then extract id
        if len(str(scenario_id)) == 4 and scenario_id is not None:

            url = f"https://pro.energytransitionmodel.com/saved_scenarios/{scenario_id}/load"
            r = requests.get(url)
            if r.status_code != 200:
                raise ValueError("Response not 200!")

            soup = BeautifulSoup(r.content, "html.parser")

            scripts = soup.find_all("script")
            for script in scripts:
                string = script.string
                if string is not None and string.find("preset_scenario_id") > 0:
                    start = (
                        string.find("preset_scenario_id")
                        + len("preset_scenario_id")
                        + 2
                    )
                    end = start + 6
                    preset_scenario_id = int(string[start:end])

            _scenario_id = preset_scenario_id
            self.locked = True  # due to using preset_scenario_id
            self.title = soup.find("span", attrs={"class": "name"}).text[1:-1]

        # if 6 digit or None then just keep that id
        elif len(str(scenario_id)) == 6 or scenario_id is None:
            _scenario_id = scenario_id
            self.locked = False
            self.title = self.get_title(_scenario_id)

        else:
            raise ValueError(f"Scenario ID provided is not valid: {scenario_id}")

        return _scenario_id


# Used by webdriver
def construct_ids(serie_or_df):
    """
    Used to combine groups, subgroups and translated name to form an ID from a workbook.
    Takes either a (sliced) df or a series (single row) to form the ID(s).

    Returns:
        list of IDs if input is pd.DataFrame
        single ID if input is pd.Series
    """
    if isinstance(serie_or_df, pd.DataFrame):
        df = serie_or_df
        ids = [
            f"{row.group} {row.subgroup} {row.translated_name}"
            for _, row in df.iterrows()
        ]

    elif isinstance(serie_or_df, pd.Series):
        row = serie_or_df
        ids = f"{row.group} {row.subgroup} {row.translated_name}"

    return ids


# Not used atm
def name2key(df):

    return lambda key: {construct_ids(row): row.key for _, row in df.iterrows()}.get(
        key
    )


def key2name(df):

    return lambda name: {row.key: construct_ids(row) for _, row in df.iterrows()}.get(
        name
    )
